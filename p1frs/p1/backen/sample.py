import cv2
import csv
import openpyxl
id= int(input("Enter student Registration number: "))
name= input("Enter student name: ")

book = openpyxl.load_workbook('Attendance\Attendance.xlsx')
sheet = book.active
max_rows = sheet.max_row+1
sheet.cell(row=max_rows,column = 1).value = id
sheet.cell(row=max_rows,column=2).value = name
book.save('Attendance\Attendance.xlsx')
def generate_dataset():
    face_classifier = cv2.CascadeClassifier("C:/Users/Shresth Jalan/Downloads/p1frs/p1/backen/haarcascade_frontalface_default.xml")
    def face_cropped(img):
        gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
        faces = face_classifier.detectMultiScale(gray, 1.3, 5)
        
        
        if faces == ():
            return None
        for (x,y,w,h) in faces:
            cropped_face = img[y:y+h,x:x+w]
        return cropped_face
    
    cap = cv2.VideoCapture(0)
    
    img_id = 0
    
    while True:
        ret, frame = cap.read()
        if face_cropped(frame) is not None:
            img_id+=1
            face = cv2.resize(face_cropped(frame), (200,200))
            face = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)
            file_name_path = "C:/Users/Shresth Jalan/Downloads/p1frs/p1/backen/data/user."+str(id)+"."+str(img_id)+".jpg"
            cv2.imwrite(file_name_path, face)
            cv2.putText(face, str(img_id), (50,50), cv2.FONT_HERSHEY_COMPLEX, 1, (0,255,0), 2)
            
            cv2.imshow("Cropped face", face)
            
        if cv2.waitKey(1)==13 or int(img_id)==50: 
            break
            
    cap.release()
    cv2.destroyAllWindows()
    print("Collecting samples is completed....")

    with open("C:/Users/Shresth Jalan/Downloads/p1frs/p1/backen/Records.csv", "w", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["Reg no.", "Name"])  # Header
        writer.writerow([id, name]) 
generate_dataset()